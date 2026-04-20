#!/usr/bin/env python3
"""Recalculate Excel formulas using LibreOffice and scan for errors.

Usage: python recalc.py <excel_file> [timeout_seconds]

Returns JSON with recalculation status and any formula errors found.
"""

import sys
import os
import json
import subprocess
import tempfile
import shutil
import re
from pathlib import Path

MACRO_NAME = "RecalcAll"
MACRO_CONTENT = """Sub RecalcAll
    Dim oDoc As Object
    Dim oSheets As Object
    Dim oSheet As Object
    Dim oCell As Object
    Dim oCursor As Object
    Dim i As Long

    oDoc = ThisComponent
    oDoc.calculateAll()
    oDoc.store()
End Sub"""

ERROR_PATTERNS = ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#GETTING_DATA"]


def find_soffice():
    """Find soffice binary."""
    script_dir = Path(__file__).parent / "office"
    soffice_wrapper = script_dir / "soffice.py"
    if soffice_wrapper.exists():
        return [sys.executable, str(soffice_wrapper)]
    for p in ["/opt/homebrew/bin/soffice", "/usr/bin/soffice", "/usr/local/bin/soffice",
              "/Applications/LibreOffice.app/Contents/MacOS/soffice"]:
        if os.path.isfile(p):
            return [p]
    soffice = shutil.which("soffice")
    if soffice:
        return [soffice]
    return None


def setup_macro():
    """Install the RecalcAll macro in LibreOffice user profile."""
    home = Path.home()
    # macOS path
    macro_dir = home / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    if not macro_dir.exists():
        # Linux path
        macro_dir = home / ".config/libreoffice/4/user/basic/Standard"

    macro_dir.mkdir(parents=True, exist_ok=True)

    module_path = macro_dir / "RecalcModule.xba"
    if not module_path.exists():
        module_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="RecalcModule" script:language="StarBasic">{MACRO_CONTENT}</script:module>"""
        module_path.write_text(module_content)

    # Register module in dialog.xlc if needed
    dialog_path = macro_dir / "dialog.xlc"
    script_path = macro_dir / "script.xlc"

    for xlc_path, tag in [(dialog_path, "dialog"), (script_path, "script")]:
        if not xlc_path.exists():
            xlc_content = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="RecalcModule"/>
</library:library>"""
            xlc_path.write_text(xlc_content)


def recalc_with_soffice(excel_path, timeout=30):
    """Recalculate using LibreOffice headless."""
    soffice_cmd = find_soffice()
    if not soffice_cmd:
        return {"status": "error", "message": "LibreOffice (soffice) not found"}

    abs_path = os.path.abspath(excel_path)

    with tempfile.TemporaryDirectory() as tmpdir:
        # Copy file to temp dir to avoid conflicts
        tmp_file = os.path.join(tmpdir, os.path.basename(abs_path))
        shutil.copy2(abs_path, tmp_file)

        # Run macro to recalculate
        cmd = soffice_cmd + [
            "--headless",
            "--calc",
            "--convert-to", "xlsx",
            "--outdir", tmpdir,
            tmp_file
        ]

        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            return {"status": "error", "message": f"LibreOffice timed out after {timeout}s"}

        # Find output file
        output_file = os.path.join(tmpdir, os.path.basename(abs_path))
        if os.path.exists(output_file):
            shutil.copy2(output_file, abs_path)

    return {"status": "recalculated"}


def scan_for_errors(excel_path):
    """Scan Excel file for formula errors using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"status": "error", "message": "openpyxl not installed"}

    wb = load_workbook(excel_path, data_only=True)
    errors = {}
    total_formulas = 0
    total_errors = 0

    # Also load with formulas to count them
    wb_formulas = load_workbook(excel_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_f = wb_formulas[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                # Check if it's a formula
                f_cell = ws_f[cell.coordinate]
                if f_cell.value and isinstance(f_cell.value, str) and f_cell.value.startswith("="):
                    total_formulas += 1

                # Check for error values
                if cell.value and isinstance(cell.value, str):
                    for err in ERROR_PATTERNS:
                        if err in str(cell.value):
                            total_errors += 1
                            if err not in errors:
                                errors[err] = {"count": 0, "locations": []}
                            errors[err]["count"] += 1
                            errors[err]["locations"].append(f"{sheet_name}!{cell.coordinate}")
                            break

    wb.close()
    wb_formulas.close()

    result = {
        "total_formulas": total_formulas,
        "total_errors": total_errors,
    }

    if errors:
        result["status"] = "errors_found"
        result["error_summary"] = errors
    else:
        result["status"] = "success"

    return result


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <excel_file> [timeout_seconds]")
        sys.exit(1)

    excel_path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    if not os.path.exists(excel_path):
        print(json.dumps({"status": "error", "message": f"File not found: {excel_path}"}))
        sys.exit(1)

    # Setup macro if needed
    setup_macro()

    # Recalculate
    recalc_result = recalc_with_soffice(excel_path, timeout)
    if recalc_result.get("status") == "error":
        print(json.dumps(recalc_result))
        sys.exit(1)

    # Scan for errors
    scan_result = scan_for_errors(excel_path)
    print(json.dumps(scan_result, indent=2))


if __name__ == "__main__":
    main()
